# -*- coding: utf-8 -*-
"""
Level 2: Excel resource. Depends on file_access.py (level 1) and
source_tracking.py (level 1) for fingerprint construction, and
openpyxl_compat.py (level 1) for warning suppression.
"""

from datetime import datetime, timezone
import warnings

from openpyxl.utils.cell import range_boundaries

import petl as etl

from task_core.excel_metadata import read_excel_row_metadata as _read_excel_row_metadata

from task_core.types import SourceCheckError
from task_core.source_tracking import (
    SourceFileMeta,
    SourceFingerprint,
    make_source_signature,
)
from task_core.file_access import _resolve_source_access, _ResourceSelection
from task_core.openpyxl_compat import suppress_openpyxl_data_validation_warning


def load_table(wb, table):
    """Extract an Excel Table's data range as a petl table. Confirmed
    correct against the original implementation this was ported from --
    a verbatim port, not a reconstruction. task_core owns it outright and
    depends on no external utility module for it."""
    ws = wb[table['sheet']]
    min_col, min_row, max_col, max_row = range_boundaries(table['range_string'])
    rows = list(ws.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col,
        values_only=True,
    ))
    header = rows[0]
    data = rows[1:]
    return etl.wrap([header, *data])


def tbl2dict(wb, table, cols=(0, 1)):
    """Build a dict from two columns of an Excel Table's data range,
    keyed by the first, deduplicated (first occurrence wins). Confirmed
    correct against the real, original implementation, same provenance
    as load_table() above."""
    ws = wb[table['sheet']]
    min_col, min_row, max_col, max_row = range_boundaries(table['range_string'])
    rows = ws.iter_rows(
        min_row=min_row + 1, max_row=max_row,  # +1: skip the header row
        min_col=min_col, max_col=max_col,
        values_only=True,
    )
    c_from, c_to = cols
    seen = set()
    return {r[c_from]: r[c_to] for r in rows if r[c_from] not in seen and not seen.add(r[c_from])}


class excel_resource:
    def __init__(self, file_path, sheets=None, tables=None, *, source_access=None, excel_buffered=False, selection=None):
        self.file_path = str(file_path)
        self._source_access = _resolve_source_access(source_access)
        self._excel_buffered = excel_buffered
        self._selection = selection
        self._workbook_cm = None
        self._wb = None
        self._table_cache = {}
        self._map_cache = {}
        self._range_cache = {}
        self._sheet_cache = {}
        self._raw_cache = {}
        # None is a safe "not yet loaded" sentinel for both: a genuinely
        # loaded .sheets is always a non-None sequence of sheet names,
        # and .tables is always a dict (possibly empty, but never None).
        # A caller may still provide both directly (the pre-existing,
        # public constructor shape -- found by external review this is
        # genuinely part of task_core's exported facade, not merely an
        # internal detail, and preserving it costs almost nothing) --
        # doing so short-circuits the lazy load entirely, the same as
        # if it had already run once.
        self._sheets_value = sheets
        self._tables_value = tables
        self._row_metadata_cache = {}

    def _ensure_metadata(self):
        # sheets and tables share one lazy load, populated together by
        # one xlsx_info() call, deliberately not split into two
        # independent lazy properties -- confirmed directly that real
        # tasks access .sheets and .tables from the same source file in
        # the same run, and splitting them would mean a second, separate
        # workbook open the first time whichever is accessed second.
        # Sharing one load keeps the existing guarantee (one open, both
        # available) exactly as it was before this became lazy at all --
        # the only thing deferred is *when* that one open happens, not
        # whether it happens.
        if self._sheets_value is None:
            with suppress_openpyxl_data_validation_warning():
                self._sheets_value, self._tables_value = self._source_access.xlsx_info(
                    self.file_path, buffered=self._excel_buffered
                )

    @property
    def sheets(self):
        self._ensure_metadata()
        return self._sheets_value

    @property
    def tables(self):
        self._ensure_metadata()
        return self._tables_value

    def source_fingerprint(self, source_key):
        if self._selection is None:
            raise SourceCheckError(
                f'excel_resource for {self.file_path!r} was not built with source-selection '
                'metadata; source_fingerprint() is only supported for resources built via '
                'build_latest_xlsx_resource()'
            )

        sel = self._selection
        f = sel.selected_file
        file_meta = SourceFileMeta(
            relative_path=f.relative_path,
            full_path=f.path,
            size_bytes=f.stat_result.st_size,
            modified_at_utc=datetime.fromtimestamp(f.stat_result.st_mtime, tz=timezone.utc),
        )

        return SourceFingerprint(
            source_key=source_key,
            source_kind=sel.source_kind,
            root_path=sel.root_path,
            include_mask=sel.include_mask,
            recursive=sel.recursive,
            file_count=1,
            total_size_bytes=file_meta.size_bytes,
            max_modified_at_utc=file_meta.modified_at_utc,
            source_signature=make_source_signature([file_meta.to_signature_dict()]),
            source_snapshot=[file_meta.to_snapshot_dict()],
            store_snapshot=True,  # file metadata, not query results -- fine to persist
        )

    def read_excel_row_metadata(self, sheet=0, mode='outline', column=None):
        # Explicit opt-in only -- never called automatically, not part of
        # build_excel_resource() or get_table(), not inferred from
        # anything. A pipeline calls this itself, from inside its own
        # .run(ctx), the same way TrackedDbQuerySource's query is the
        # task author's decision, not scaffold inference. Returns the
        # raw {row_number: value} mapping; alignment against a specific
        # materialized table's row range (task_core.align_row_metadata,
        # in excel_metadata.py) stays the pipeline author's own call,
        # since only the pipeline knows how its own table's rows
        # correspond to XLSX row numbers after any filtering it does of
        # its own.
        #
        # Cached by (sheet, mode, column), not left to the caller: this
        # resource represents one immutable selected workbook, so the
        # same key always means the same answer, and caching it here
        # removes a real, if easy to forget, cost -- opening its own
        # handle via open_binary(), independent of whatever handle this
        # resource may already be holding open for its normal data reads
        # (in SMB non-buffered mode, that handle stays open for the
        # resource's whole lifetime), is a second full network read on
        # every single call otherwise, not just the first.
        key = (sheet, mode, column)
        if key not in self._row_metadata_cache:
            with self._source_access.open_binary(self.file_path, buffered=self._excel_buffered) as src:
                self._row_metadata_cache[key] = _read_excel_row_metadata(src, sheet=sheet, mode=mode, column=column)
        # A copy, not the internal cache dict directly -- found by a
        # further review, confirmed directly: a caller mutating what
        # this returned silently corrupted the cached answer for a
        # later pipeline calling this with the same key. The same class
        # of regression already fixed for get_sheet_raw_rows() (see its
        # own docstring), applied here for the same reason.
        return dict(self._row_metadata_cache[key])

    def _ensure_workbook(self):
        if self._wb is None:
            # Metadata resolved -- opened, closed, GC'd -- before the
            # retained workbook below, regardless of which method got
            # here first. Found by a further review: the earlier fix
            # (self.tables[name] evaluated before _ensure_workbook() in
            # get_table()/get_map() specifically) only prevented THOSE
            # two methods from triggering the overlap themselves -- it
            # didn't protect against _ensure_workbook() already having
            # been triggered by some OTHER, earlier call (get_sheet_rows(),
            # get_range(), etc.) before get_table() ever runs. Confirmed
            # directly with the same source_access that rejects
            # xlsx_info() while a retained workbook is active: calling
            # get_sheet_rows() then get_table() still overlapped, even
            # with the earlier fix in place. This is the one, shared
            # choke point every method that needs the retained workbook
            # goes through, so resolving metadata here protects every
            # call path, not just the two originally touched.
            self._ensure_metadata()
            # We intentionally keep the workbook context open across the whole
            # excel_resource lifetime so repeated get_table/get_map/get_range/
            # get_sheet_rows calls reuse one workbook instance and one set of
            # caches.
            #
            # In SMB non-buffered mode this also means the remote file stream
            # stays open until excel_resource.close(). In buffered mode the SMB
            # stream is closed immediately after reading into memory; only the
            # in-memory buffer remains open for the workbook lifetime.
            with suppress_openpyxl_data_validation_warning():
                self._workbook_cm = self._source_access.open_workbook(
                    self.file_path,
                    read_only=True,
                    data_only=True,
                    buffered=self._excel_buffered,
                )
                self._wb = self._workbook_cm.__enter__()
        return self._wb

    def get_table(self, name):
        if name not in self._table_cache:
            with suppress_openpyxl_data_validation_warning():
                # _ensure_workbook() itself now guarantees metadata is
                # resolved before the retained workbook opens (see its
                # own comment), regardless of call order -- this method
                # no longer needs to enforce that ordering itself.
                wb = self._ensure_workbook()
                self._table_cache[name] = load_table(wb, self.tables[name])
        return self._table_cache[name]

    def get_map(self, name, cols=(0, 1)):
        key = (name, cols)
        if key not in self._map_cache:
            with suppress_openpyxl_data_validation_warning():
                wb = self._ensure_workbook()
                self._map_cache[key] = tbl2dict(wb, self.tables[name], cols=cols)
        return self._map_cache[key]

    def get_range(self, sheet, range_string, header=True):
        # header=True/False produce an identical result -- confirmed
        # directly, not assumed. A petl table's first row is always
        # treated as its header by etl.header()/etl.data() regardless of
        # how the table was constructed, so there is no way for this
        # method's return value (always a petl table) to represent
        # "headerless" at all. Use get_sheet_raw_rows() instead for a
        # genuinely header-agnostic plain row sequence. Kept as its own
        # method rather than changed in place: no real task in this
        # project calls this with header=False anymore (hr_task.py's own
        # former caller, sheet_to_raw_dataframe(), was migrated to
        # get_sheet_raw_rows() instead), but an external caller relying
        # on the old, documented-as-accidental behavior might still exist.
        if header is False:
            warnings.warn(
                "get_range(header=False) has no effect -- a petl table's first row is always "
                "treated as its header by etl.header()/etl.data() regardless of this argument. "
                "Use get_sheet_raw_rows() for genuinely header-agnostic rows.",
                DeprecationWarning,
                stacklevel=2,
            )
        key = (sheet, range_string, header)
        if key not in self._range_cache:
            with suppress_openpyxl_data_validation_warning():
                wb = self._ensure_workbook()
                ws = wb[sheet]
                min_col, min_row, max_col, max_row = range_boundaries(range_string)
                rows = list(
                    ws.iter_rows(
                        min_row=min_row,
                        max_row=max_row,
                        min_col=min_col,
                        max_col=max_col,
                        values_only=True,
                    )
                )
                if not rows:
                    self._range_cache[key] = etl.empty()
                elif header:
                    self._range_cache[key] = etl.wrap([rows[0], *rows[1:]])
                else:
                    self._range_cache[key] = etl.wrap(rows)
        return self._range_cache[key]

    def get_sheet_rows(self, sheet, header=True):
        # Same header=True/False no-op as get_range() above, same reason,
        # same fix (get_sheet_raw_rows() below). No real task in this
        # project calls this with header=False anymore -- hr_task.py's
        # own former caller, sheet_to_raw_dataframe(), was migrated to
        # get_sheet_raw_rows() instead.
        if header is False:
            warnings.warn(
                "get_sheet_rows(header=False) has no effect -- a petl table's first row is always "
                "treated as its header by etl.header()/etl.data() regardless of this argument. "
                "Use get_sheet_raw_rows() for genuinely header-agnostic rows.",
                DeprecationWarning,
                stacklevel=2,
            )
        # Shares _raw_rows_for_sheet()'s own cache rather than reading
        # the worksheet a second time -- both previously called the
        # identical list(ws.values) independently, genuinely storing the
        # same sheet's rows twice if both were ever used for it. Cache
        # key is sheet alone now, not (sheet, header): header never
        # affected the result (the confirmed no-op above), so the two
        # were already producing identical values as two separate,
        # non-identical cache entries -- now genuinely one. Built from
        # the PRIVATE, internal materialization (never exposed directly
        # to an external caller), not get_sheet_raw_rows()'s own public
        # return value -- see that method's own docstring for why this
        # distinction matters.
        rows = self._raw_rows_for_sheet(sheet)
        if sheet not in self._sheet_cache:
            self._sheet_cache[sheet] = etl.empty() if not rows else etl.wrap(rows)
        return self._sheet_cache[sheet]

    def _raw_rows_for_sheet(self, sheet):
        # The one, genuine list(ws.values) read for a given sheet,
        # cached and shared internally between get_sheet_rows() and
        # get_sheet_raw_rows() -- never returned directly to an external
        # caller. See get_sheet_raw_rows()'s own docstring for why this
        # needs to stay separate from what that method actually returns.
        if sheet not in self._raw_cache:
            with suppress_openpyxl_data_validation_warning():
                wb = self._ensure_workbook()
                ws = wb[sheet]
                self._raw_cache[sheet] = list(ws.values)
        return self._raw_cache[sheet]

    def get_sheet_raw_rows(self, sheet):
        """Every row of the sheet, as a plain list of tuples -- no
        header concept at all, unlike get_sheet_rows()/get_range(),
        whose return value is always a petl table and therefore always
        treats its own first row as a header once read via
        etl.header()/etl.data(). Use this when row 0 needs to stay
        available as genuine data (e.g. scanning for a header row that
        could be anywhere in the sheet), instead of reconstructing it
        via etl.header()/etl.data() on a get_sheet_rows(header=False)
        result.

        Returns a fresh copy every call, not the resource's own internal
        cache directly -- found by a further review, confirmed directly:
        after get_sheet_rows()/get_sheet_raw_rows() were made to share
        one materialization, mutating what this method returned (e.g.
        raw[1] = (9, 9)) silently changed get_sheet_rows()'s own,
        already-cached petl table too, since both were built from the
        exact same, single list object. The underlying list(ws.values)
        read still only happens once per sheet, cached internally
        (_raw_rows_for_sheet() above) -- only the copy on the way out is
        new, a cheap, shallow list() copy, negligible next to the actual
        costs (SMB reads, workbook opens) this project's caching has
        been about."""
        return list(self._raw_rows_for_sheet(sheet))

    def close(self):
        # References dropped BEFORE __exit__(), not after. open_workbook()'s
        # own finally: now runs gc.collect() from inside this __exit__() call
        # -- and self._wb still pointing at the workbook at that moment would
        # defeat it exactly the way a skipped `del wb` defeats xlsx_info()'s.
        # Confirmed directly, before this change: at the instant __exit__()
        # returned, gc.get_referrers() on the workbook still listed this
        # resource's own __dict__, because _wb was only cleared afterwards in
        # the finally: below.
        #
        # Swap-then-close, so both attributes are cleared even if __exit__()
        # raises -- the same shape DbPublisher.close() already uses, and the
        # reason a second close() after a failed one is a clean no-op rather
        # than a retry of the same failing workbook.
        workbook_cm, self._workbook_cm = self._workbook_cm, None
        self._wb = None

        try:
            if workbook_cm is not None:
                # open_workbook()'s own context manager (file_access.py)
                # already does `finally: wb.close()` -- this __exit__()
                # call closes the workbook; no separate, explicit
                # self._wb.close() is needed alongside it.
                workbook_cm.__exit__(None, None, None)
        finally:
            # Confirmed directly that none of these caches transitively
            # reference the workbook or any worksheet (they hold plain
            # values, tuples and petl tables built from them), so clearing
            # them after the collect above costs the collect nothing. They
            # are cleared in a finally: regardless, so a failing __exit__()
            # cannot leave this resource holding stale cached data for a
            # workbook that is gone.
            self._table_cache.clear()
            self._map_cache.clear()
            self._range_cache.clear()
            self._sheet_cache.clear()
            self._raw_cache.clear()
            self._row_metadata_cache.clear()



def build_excel_resource(file_path, *, source_access=None, excel_buffered=False, selection=None):
    source_access = _resolve_source_access(source_access)
    file_path = source_access.select_fixed_file(file_path)

    return excel_resource(
        file_path=file_path,
        source_access=source_access,
        excel_buffered=excel_buffered,
        selection=selection,
    )



def build_latest_xlsx_resource(
    folder_path,
    pattern='*.xlsx',
    *,
    include_hidden=False,
    include_system=False,
    include_temp=False,
    min_age_seconds=None,
    recursive=False,
    source_access=None,
    excel_buffered=False,
):
    source_access = _resolve_source_access(source_access)

    # Select via select_file_infos() (not select_latest_file()) so the
    # SelectedFile/stat_result used for the source_fingerprint() below is the
    # same one that picks the file to open -- no second directory listing or
    # re-stat.
    file_infos = source_access.select_file_infos(
        folder_path,
        pattern=pattern,
        include_hidden=include_hidden,
        include_system=include_system,
        include_temp=include_temp,
        min_age_seconds=min_age_seconds,
        recursive=recursive,
    )
    latest = max(file_infos, key=lambda item: (item.stat_result.st_mtime, item.path))

    selection = _ResourceSelection(
        source_kind='latest_file',
        root_path=str(folder_path),
        include_mask=pattern,
        recursive=recursive,
        selected_file=latest,
    )

    return build_excel_resource(
        latest.path,
        source_access=source_access,
        excel_buffered=excel_buffered,
        selection=selection,
    )
