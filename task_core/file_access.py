"""
Level 1: local/SMB file selection and Excel-workbook-metadata access.
Depends on openpyxl_compat.py, a same-level (level 1) lateral dependency
for warning suppression -- does not need types.py or source_tracking.py
at all.

_table_ref() lives here, not in resources/excel.py, deliberately: it's
called from inside xlsx_info(), a source_access method -- putting it in
resources/excel.py (level 2) would make this module (level 1) import
upward, which the package's own layering rule forbids.
"""

import gc
import io
import stat
import time
from contextlib import contextmanager
from dataclasses import dataclass
from fnmatch import fnmatch
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from task_core.openpyxl_compat import suppress_openpyxl_data_validation_warning


@dataclass(frozen=True)
class SelectedFile:
    path: str
    relative_path: str
    stat_result: Any


class NoMatchingFilesError(FileNotFoundError):
    """The folder exists, but the pattern matched zero files -- distinct
    from the folder itself not existing, which stays a plain
    FileNotFoundError. Subclassing FileNotFoundError means any existing
    broad `except FileNotFoundError:` still catches this too; only new
    code that needs the distinction has to ask for it (see
    resources/file_set.py's on_empty='empty' handling)."""



@dataclass(frozen=True)
class _ResourceSelection:
    # Internal: records how a resource's underlying file(s) were selected,
    # so the resource can later answer source_fingerprint() using the same
    # selection rules the actual data load used (see SMB/DFS support notes:
    # the fingerprint must reflect the same file set that gets processed).
    source_kind: str
    root_path: str | None
    include_mask: str | None
    recursive: bool
    selected_file: SelectedFile


def _has_windows_attr(path, attr_flag):
    try:
        return bool(path.stat().st_file_attributes & attr_flag)
    except AttributeError:
        return False



# Compatibility surface -- confirmed unused anywhere in this project,
# even internally (grep found nothing beyond these definitions
# themselves). The real filtering loop in _select_local_file_infos below
# uses its own, private _has_windows_attr_from_stat() directly, not
# these. Not removed without checking for external consumers first.


def is_hidden_file(path):
    path = Path(path)

    if path.name.startswith('.'):
        return True

    return _has_windows_attr(path, stat.FILE_ATTRIBUTE_HIDDEN)


def is_system_file(path):
    path = Path(path)
    return _has_windows_attr(path, stat.FILE_ATTRIBUTE_SYSTEM)



def is_excel_temp_file(path):
    path = Path(path)
    return path.name.startswith('~$')


def _table_ref(tbl):
    # Older openpyxl table objects are themselves range-string-like; newer
    # ones expose .ref. Isolated as a function so xlsx_info() can use it
    # from within a comprehension.
    try:
        return tbl.ref
    except AttributeError:
        return tbl


class source_access:
    """Local or SMB file and workbook access.

    Named `file_access` until 0.3.1. The facade re-exported a class with
    its own module's name, so `task_core.file_access` resolved to the
    CLASS and shadowed the submodule -- the same trap as types.py
    shadowing the stdlib, and it cost a working session a wrong turn.
    `source_access` matches build_source_access(), which is how every
    caller actually obtains one; nothing imported the class by name.
    """

    def __init__(self, dfs_creds=None):
        self._dfs_creds = dfs_creds
        self._conn_kwargs = {}

        if dfs_creds is not None:
            self._conn_kwargs = {
                'username': dfs_creds['username'],
                'password': dfs_creds['password'],
                'port': dfs_creds.get('port', 445),
            }

    def _get_smbclient(self):
        import smbclient
        return smbclient

    def _is_unc_path(self, path):
        return str(path).startswith('\\\\')

    def _use_smb(self, path):
        return self._dfs_creds is not None and self._is_unc_path(path)

    def _stat(self, path):
        # Coerced to str unconditionally, local branch and SMB branch
        # alike -- a Path object works fine almost everywhere (pathlib
        # methods accept it transparently), but smbclient's own internal
        # DFS-referral resolution does raw string operations (slicing,
        # .split('\\')) on whatever it's given, and breaks on a Path
        # object with a real, hard-to-diagnose TypeError, but only on
        # genuine DFS referrals -- easy to miss in testing against a
        # plain SMB share. Confirmed directly: hr_task.py hit exactly
        # this, passing a WindowsPath into what became this call.
        path = str(path)
        if not self._use_smb(path):
            return Path(path).stat()
        return self._get_smbclient().stat(path, **self._conn_kwargs)

    def _has_windows_attr_from_stat(self, st, attr_flag):
        try:
            return bool(st.st_file_attributes & attr_flag)
        except AttributeError:
            return False

    def _is_hidden_remote(self, name, st):
        if name.startswith('.'):
            return True
        return self._has_windows_attr_from_stat(st, stat.FILE_ATTRIBUTE_HIDDEN)

    def _is_system_remote(self, st):
        return self._has_windows_attr_from_stat(st, stat.FILE_ATTRIBUTE_SYSTEM)

    def _matches_pattern(self, name, rel_path, pattern):
        # Match both basename and relative path so callers can use either
        # simple file masks ('*.xlsx') or folder-aware masks ('subdir/*.xlsx').
        rel_path_norm = rel_path.replace('\\', '/')
        pattern_norm = pattern.replace('\\', '/')
        return fnmatch(name, pattern) or fnmatch(rel_path_norm, pattern_norm)

    def select_fixed_file(self, path):
        path = str(path)  # same reasoning as _stat/open_binary/select_file_infos
        if not self._use_smb(path):
            path = Path(path)
            if not path.exists():
                raise FileNotFoundError(f'File not found: {path}')
            if not path.is_file():
                raise FileNotFoundError(f'Path is not a file: {path}')
            return str(path)

        try:
            st = self._stat(path)
        except FileNotFoundError as e:
            raise FileNotFoundError(f'File not found: {path}') from e

        if not stat.S_ISREG(st.st_mode):
            raise FileNotFoundError(f'Path is not a file: {path}')

        return str(path)

    def select_file_infos(
        self,
        folder_path,
        pattern='*',
        *,
        include_hidden=False,
        include_system=False,
        include_temp=False,
        min_age_seconds=None,
        recursive=False,
    ):
        folder_path = str(folder_path)  # same reasoning as _stat/open_binary
                                         # -- coerced once here, at the public
                                         # entry point, propagates to both
                                         # _select_local_file_infos and
                                         # _select_smb_file_infos below, the
                                         # only two callers of either.
        if not self._use_smb(folder_path):
            return self._select_local_file_infos(
                folder_path,
                pattern=pattern,
                include_hidden=include_hidden,
                include_system=include_system,
                include_temp=include_temp,
                min_age_seconds=min_age_seconds,
                recursive=recursive,
            )

        return self._select_smb_file_infos(
            folder_path,
            pattern=pattern,
            include_hidden=include_hidden,
            include_system=include_system,
            include_temp=include_temp,
            min_age_seconds=min_age_seconds,
            recursive=recursive,
        )

    def _select_local_file_infos(
        self,
        folder_path,
        *,
        pattern,
        include_hidden,
        include_system,
        include_temp,
        min_age_seconds,
        recursive,
    ):
        folder = Path(folder_path)

        if not folder.exists():
            raise FileNotFoundError(f'Path not found: {folder}')

        if folder.is_file():
            raise ValueError(
                f'select_file_infos() is for scanning a folder, not selecting a specific file: '
                f'{folder} is a file, not a directory. Use select_fixed_file() instead -- '
                f'select_file_infos() applying its filters (pattern, hidden/system/temp, '
                f'min_age_seconds) to an explicitly-named file would be a meaningless, '
                f'surprising contract either way (skip them all, or reject a file the caller '
                f'named directly because it happens not to match a glob pattern).'
            )

        iterator = folder.rglob(pattern) if recursive else folder.glob(pattern)
        now = time.time()
        files = []

        for fn in iterator:
            try:
                st = fn.stat()
            except FileNotFoundError:
                continue

            if not stat.S_ISREG(st.st_mode):
                continue

            if not include_temp and is_excel_temp_file(fn):
                continue

            if not include_hidden and (fn.name.startswith('.') or self._has_windows_attr_from_stat(st, stat.FILE_ATTRIBUTE_HIDDEN)):
                continue

            if not include_system and self._has_windows_attr_from_stat(st, stat.FILE_ATTRIBUTE_SYSTEM):
                continue

            if min_age_seconds is not None:
                age_seconds = now - st.st_mtime
                if age_seconds < float(min_age_seconds):
                    continue

            try:
                relative_path = str(fn.relative_to(folder))
            except ValueError:
                relative_path = fn.name

            files.append(SelectedFile(path=str(fn), relative_path=relative_path, stat_result=st))

        if not files:
            raise NoMatchingFilesError(f'No files found in {folder} for pattern {pattern!r}')

        return files

    def _select_smb_file_infos(
        self,
        folder_path,
        *,
        pattern,
        include_hidden,
        include_system,
        include_temp,
        min_age_seconds,
        recursive,
    ):
        try:
            root_st = self._stat(folder_path)
        except FileNotFoundError as e:
            raise FileNotFoundError(f'Path not found: {folder_path}') from e

        if stat.S_ISREG(root_st.st_mode):
            raise ValueError(
                f'select_file_infos() is for scanning a folder, not selecting a specific file: '
                f'{folder_path} is a file, not a directory. Use select_fixed_file() instead.'
            )

        if not stat.S_ISDIR(root_st.st_mode):
            raise FileNotFoundError(f'Path not found: {folder_path}')

        smbclient = self._get_smbclient()
        now = time.time()
        files = []
        base = str(folder_path).rstrip('\\/')

        for dirpath, dirnames, filenames in smbclient.walk(base, **self._conn_kwargs):
            rel_dir = dirpath[len(base):].lstrip('\\/') if dirpath != base else ''

            if not recursive:
                dirnames[:] = []

            for filename in filenames:
                rel_path = filename if not rel_dir else rel_dir + '\\' + filename

                if not self._matches_pattern(filename, rel_path, pattern):
                    continue

                full_path = dirpath.rstrip('\\/') + '\\' + filename if dirpath.rstrip('\\/') else filename
                st = self._stat(full_path)

                if not stat.S_ISREG(st.st_mode):
                    continue

                if not include_temp and filename.startswith('~$'):
                    continue

                if not include_hidden and self._is_hidden_remote(filename, st):
                    continue

                if not include_system and self._is_system_remote(st):
                    continue

                if min_age_seconds is not None:
                    age_seconds = now - st.st_mtime
                    if age_seconds < float(min_age_seconds):
                        continue

                files.append(SelectedFile(path=full_path, relative_path=rel_path, stat_result=st))

        if not files:
            raise NoMatchingFilesError(f'No files found in {folder_path} for pattern {pattern!r}')

        return files

    def select_files(
        self,
        folder_path,
        pattern='*',
        *,
        include_hidden=False,
        include_system=False,
        include_temp=False,
        min_age_seconds=None,
        recursive=False,
    ):
        return [
            item.path
            for item in self.select_file_infos(
                folder_path,
                pattern=pattern,
                include_hidden=include_hidden,
                include_system=include_system,
                include_temp=include_temp,
                min_age_seconds=min_age_seconds,
                recursive=recursive,
            )
        ]

    def select_latest_file(
        self,
        folder_path,
        pattern='*',
        *,
        include_hidden=False,
        include_system=False,
        include_temp=False,
        min_age_seconds=None,
        recursive=False,
    ):
        files = self.select_file_infos(
            folder_path,
            pattern=pattern,
            include_hidden=include_hidden,
            include_system=include_system,
            include_temp=include_temp,
            min_age_seconds=min_age_seconds,
            recursive=recursive,
        )
        latest = max(files, key=lambda item: (item.stat_result.st_mtime, item.path))
        return latest.path

    @contextmanager
    def open_binary(self, path, *, buffered=False):
        path = str(path)  # same reasoning as _stat -- smbclient.open_file()
                           # would hit the identical DFS-referral crash on a
                           # Path object, and this is the other place a raw
                           # path reaches smbclient directly.
        if not self._use_smb(path):
            with open(path, 'rb') as f:
                yield f
            return

        smbclient = self._get_smbclient()

        with smbclient.open_file(path, mode='rb', **self._conn_kwargs) as f:
            if buffered:
                buf = io.BytesIO(f.read())
                try:
                    yield buf
                finally:
                    buf.close()
            else:
                yield f

    @contextmanager
    def open_workbook(self, path, *, buffered=False, **kwargs):
        # Same del wb + gc.collect() treatment as xlsx_info() below, and for
        # a stronger reason than xlsx_info() has: this is the workbook
        # excel_resource RETAINS for its whole lifetime (see
        # _ensure_workbook()), so in SMB non-buffered mode the remote stream
        # stays open until close(). xlsx_info()'s workbook lives for
        # milliseconds; this one can live for the entire task. If openpyxl
        # object cycles genuinely keep XLSX/ZIP handles open on SMB/DFS after
        # an explicit close() -- the confirmed production finding
        # gc.collect() exists here for at all -- then this path is where
        # that costs the most, and it previously had no gc.collect() on it
        # whatsoever. Confirmed directly by counting calls: xlsx_info() made
        # exactly 1, open_workbook() and excel_resource.close() made 0.
        #
        # Structure mirrors xlsx_info() deliberately rather than
        # approximately: del wb in its OWN inner finally, so a raising
        # wb.close() cannot skip it and leave wb a live local at collect
        # time; and gc.collect() in an OUTER finally, so it runs whether
        # load_workbook() succeeded, the body raised, or wb.close() itself
        # raised -- and so that it runs only after open_binary()'s own
        # context has exited, never while the underlying stream could still
        # hold a reference.
        try:
            with self.open_binary(path, buffered=buffered) as src:
                wb = load_workbook(src, **kwargs)
                # The try: starts AFTER the assignment, so a failing
                # load_workbook() never reaches `del wb` with wb unbound.
                try:
                    yield wb
                finally:
                    try:
                        wb.close()
                    finally:
                        del wb
        finally:
            gc.collect()

    def xlsx_info(self, path, tables=True, *, buffered=False):
        with suppress_openpyxl_data_validation_warning():
            try:
                with self.open_binary(path, buffered=buffered) as src:
                    wb = load_workbook(
                        src,
                        # When only sheet names are needed we can use read_only=True
                        # for a lighter/faster open. Named table metadata lives on
                        # ws.tables, which requires normal workbook mode.
                        read_only=not tables,
                        data_only=True,
                        keep_links=False,
                    )
                    try:
                        sheets = wb.sheetnames
                        t = {
                            tbl_name: {'sheet': sht_name, 'range_string': _table_ref(tbl)}
                            for sht_name in sheets
                            for tbl_name, tbl in wb[sht_name].tables.items()
                        } if tables else {}
                    finally:
                        # No explicit `del ws` needed here (unlike the old
                        # for-loop version): comprehension variables don't leak
                        # into this scope, so each worksheet reference is
                        # already dropped as soon as the next iteration starts --
                        # at least as prompt as the old explicit del, often more so.
                        #
                        # del wb in its own, inner finally -- found by a
                        # further review, confirmed directly: if
                        # wb.close() itself raised, del wb was skipped
                        # entirely (an exception from one statement in a
                        # finally: block skips the rest of that same
                        # block), leaving wb as a live, reachable local
                        # variable at the point the outer gc.collect()
                        # below runs -- weakening the exact stuck-handle
                        # workaround it exists to guarantee, on the one
                        # path (a failing close()) where that guarantee
                        # matters most.
                        try:
                            wb.close()
                        finally:
                            del wb
            finally:
                # Required in production: openpyxl object cycles can otherwise
                # keep XLSX/ZIP handles open on SMB/DFS after explicit close().
                # This outer try/finally, not just the inner workbook-close one
                # above, so gc.collect() runs on every path, not only success --
                # a failure during metadata scanning (load_workbook() itself
                # raising, or the table-metadata comprehension above it) still
                # risks the same stuck-handle behavior this exists to prevent,
                # and previously skipped gc.collect() entirely in that case.
                # Deliberately still outside the inner workbook-close finally,
                # not merged into it: gc.collect() runs once, after the
                # workbook is closed and the underlying stream itself has
                # also exited via open_binary()'s own __exit__, not any
                # earlier, while either could still be holding a reference.
                gc.collect()

        return sheets, t


LOCAL_FILE_ACCESS = source_access()


def build_source_access(dfs_creds=None):
    return LOCAL_FILE_ACCESS if dfs_creds is None else source_access(dfs_creds=dfs_creds)


def _resolve_source_access(source_access=None):
    return LOCAL_FILE_ACCESS if source_access is None else source_access


def select_fixed_file(path, *, source_access=None):
    return _resolve_source_access(source_access).select_fixed_file(path)



def select_file_infos(
    folder_path,
    pattern='*',
    *,
    include_hidden=False,
    include_system=False,
    include_temp=False,
    min_age_seconds=None,
    recursive=False,
    source_access=None,
):
    return _resolve_source_access(source_access).select_file_infos(
        folder_path,
        pattern=pattern,
        include_hidden=include_hidden,
        include_system=include_system,
        include_temp=include_temp,
        min_age_seconds=min_age_seconds,
        recursive=recursive,
    )


# Compatibility surface -- confirmed unused anywhere in this project
# (grep found nothing beyond these definitions and their own calls into
# the underlying class methods). Every real caller uses select_file_infos()
# directly instead. Not removed without checking for external consumers
# first.


def select_files(
    folder_path,
    pattern='*',
    *,
    include_hidden=False,
    include_system=False,
    include_temp=False,
    min_age_seconds=None,
    recursive=False,
    source_access=None,
):
    return _resolve_source_access(source_access).select_files(
        folder_path,
        pattern=pattern,
        include_hidden=include_hidden,
        include_system=include_system,
        include_temp=include_temp,
        min_age_seconds=min_age_seconds,
        recursive=recursive,
    )



def select_latest_file(
    folder_path,
    pattern='*',
    *,
    include_hidden=False,
    include_system=False,
    include_temp=False,
    min_age_seconds=None,
    recursive=False,
    source_access=None,
):
    return _resolve_source_access(source_access).select_latest_file(
        folder_path,
        pattern=pattern,
        include_hidden=include_hidden,
        include_system=include_system,
        include_temp=include_temp,
        min_age_seconds=min_age_seconds,
        recursive=recursive,
    )
