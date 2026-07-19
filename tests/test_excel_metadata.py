# -*- coding: utf-8 -*-
"""
Real, generated .xlsx fixtures for excel_metadata.py -- mocks alone
aren't sufficient here, since the behavior under test is genuinely
XML-specific (sparse rows, outline levels, sheet relationships) that a
mock would have to reimplement correctly to be worth anything, at which
point it's not testing the real code anymore.

Test 2 exists because this exact class of bug -- first_row misalignment
against a genuinely untouched leading row (no <row r="1"> in the raw XML
at all) -- was found twice in this project's history through manual,
ad hoc verification (once in hr_task.py's read_ssch_sheet, again
independently in hr_petl_task.py's read_ssch2_sheet), never as a
persistent, automated test until now.
"""

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from task_core.excel_metadata import align_row_metadata, get_sheets, read_excel_row_metadata


class Test1MultipleSheetsAndNameResolution(unittest.TestCase):
    def setUp(self):
        self._tmpdir = tempfile.TemporaryDirectory()
        self.path = Path(self._tmpdir.name) / 'multi_sheet.xlsx'

        wb = Workbook()
        wb.active.title = 'First'
        wb.active['A1'] = 'first-sheet-marker'
        wb.active.row_dimensions[1].outline_level = 0

        second = wb.create_sheet('Second')
        second['A1'] = 'second-sheet-marker'
        second.row_dimensions[1].outline_level = 2  # distinct from every other sheet

        third = wb.create_sheet('Third')
        third['A1'] = 'third-sheet-marker'
        third.row_dimensions[1].outline_level = 0

        wb.save(self.path)

    def tearDown(self):
        self._tmpdir.cleanup()

    def test_get_sheets_returns_names_in_order(self):
        import zipfile
        with zipfile.ZipFile(self.path) as zf:
            self.assertEqual(get_sheets(zf), ['First', 'Second', 'Third'])

    def test_read_by_name_resolves_the_correct_sheet(self):
        # 'Second' is the only sheet with outline_level=2 -- distinctive
        # enough that reading the wrong sheet would fail this, unlike
        # asserting a shape every sheet happens to share.
        result = read_excel_row_metadata(str(self.path), sheet='Second', mode='outline')
        self.assertEqual(result, {1: 2})

    def test_unknown_sheet_name_raises_clearly(self):
        with self.assertRaisesRegex(KeyError, 'Sheet name not found'):
            read_excel_row_metadata(str(self.path), sheet='DoesNotExist', mode='outline')


class Test2FirstRowAlignmentAgainstGenuinelySparseXml(unittest.TestCase):
    """The exact construction that caught the first_row bug, twice, in
    this project's history -- captured here permanently so a regression
    is caught automatically, not just if someone happens to build this
    same fixture by hand again."""

    def setUp(self):
        self._tmpdir = tempfile.TemporaryDirectory()
        self.path = Path(self._tmpdir.name) / 'sparse_leading_row.xlsx'

        wb = Workbook()
        ws = wb.active
        # Row 1 is genuinely untouched -- no cell written to it at all, so
        # openpyxl never emits a <row r="1"> element in the raw XML.
        # Row 2 onward are real, giving metadata = {2: 0, 4: 0, 8: 1}
        # (rows without any content in between get no XML entry either).
        ws.cell(row=2, column=1, value='marker')
        ws.cell(row=4, column=2, value='header')
        ws.cell(row=8, column=2, value='Target Row')
        ws.row_dimensions[8].outline_level = 1
        wb.save(self.path)

    def tearDown(self):
        self._tmpdir.cleanup()

    def test_metadata_has_gaps_where_xml_rows_are_genuinely_absent(self):
        metadata = read_excel_row_metadata(str(self.path), sheet=0, mode='outline')
        # Row 1, 3, 5, 6, 7 never got a <row> element at all -- confirming
        # the fixture is genuinely sparse, not just visually sparse.
        self.assertNotIn(1, metadata)
        self.assertEqual(metadata.get(8), 1)

    def test_first_row_equals_one_aligns_correctly_despite_the_gap(self):
        # 8 physical rows exist once materialized (e.g. via pandas/petl,
        # which always produce a row for every position up to the last
        # used row, regardless of XML sparseness) -- row 0 in that
        # materialization is XLSX row 1, even though row 1 has no XML
        # entry of its own.
        metadata = read_excel_row_metadata(str(self.path), sheet=0, mode='outline')
        aligned = align_row_metadata(metadata, first_row=1, n_rows=8)

        # Row 8 (index 7) is 'Target Row', outline_level=1 -- this is the
        # exact assertion that catches the regression: with the old,
        # buggy `first_row = min(metadata)` (= 2, not 1), this value
        # would land one position earlier, at index 6, and index 7 would
        # incorrectly be None.
        self.assertEqual(aligned[7], 1, "Target Row's outline_level did not align to the correct position")

    def test_min_metadata_as_first_row_would_have_been_wrong(self):
        # Documents *why* first_row=1 is correct and not incidental --
        # the old, buggy approach used min(metadata) as first_row, which
        # for this exact fixture would be 2, not 1, silently shifting
        # every position by one.
        metadata = read_excel_row_metadata(str(self.path), sheet=0, mode='outline')
        buggy_first_row = min(metadata)
        self.assertEqual(buggy_first_row, 2, 'fixture no longer reproduces the historical bug scenario')
        buggy_aligned = align_row_metadata(metadata, first_row=buggy_first_row, n_rows=8)
        self.assertIsNone(buggy_aligned[7], 'the buggy alignment should misplace the target row -- if this fails, the fixture stopped exercising the bug')


class Test3OutlineLevelExtraction(unittest.TestCase):
    def setUp(self):
        self._tmpdir = tempfile.TemporaryDirectory()
        self.path = Path(self._tmpdir.name) / 'outline_levels.xlsx'

        wb = Workbook()
        ws = wb.active
        for row, level in [(1, 0), (2, 1), (3, 2), (4, 1), (5, 0)]:
            ws.cell(row=row, column=1, value=f'row-{row}')
            ws.row_dimensions[row].outline_level = level
        wb.save(self.path)

    def tearDown(self):
        self._tmpdir.cleanup()

    def test_each_rows_outline_level_extracted_correctly(self):
        metadata = read_excel_row_metadata(str(self.path), sheet=0, mode='outline')
        self.assertEqual(metadata, {1: 0, 2: 1, 3: 2, 4: 1, 5: 0})


if __name__ == '__main__':
    unittest.main()
