# -*- coding: utf-8 -*-
"""
File-resource selection and fingerprinting -- which physical input gets
processed, and whether a fingerprint change correctly triggers a rerun.
Real temp directories and real files throughout, not mocks: filesystem
timing, ordering, and stat behavior are exactly the kind of thing a mock
would have to reimplement correctly to be worth testing at all.
"""

import os
import tempfile
import time
import unittest
from pathlib import Path

import task_core as tc
import petl as etl
from task_core.file_access import source_access as FileAccessImpl, NoMatchingFilesError
from task_core.resources.excel import build_latest_xlsx_resource
from task_core.resources.file_set import build_file_set_resource


class TempDir:
    def __enter__(self):
        self._tmpdir = tempfile.TemporaryDirectory()
        return Path(self._tmpdir.name)

    def __exit__(self, *exc):
        self._tmpdir.cleanup()


def _write(path, content=b'x'):
    path.write_bytes(content)


def _write_xlsx(path):
    from openpyxl import Workbook
    Workbook().save(path)


class _FixedOrderAccess:
    """Wraps select_file_infos to hand results back in a caller-specified
    order, regardless of whatever order the real filesystem's glob
    happens to produce. Needed because these tests must genuinely prove
    mtime-based selection, not pass by coincidence of glob's own,
    filesystem-dependent order -- found by deliberately breaking
    build_latest_xlsx_resource's selection logic entirely and discovering
    the first version of these tests still passed, since glob on this
    filesystem happened to return the newer file first anyway."""

    def __init__(self, order):
        self._real = tc.LOCAL_FILE_ACCESS
        self._order = order

    def select_file_infos(self, *a, **kw):
        infos = self._real.select_file_infos(*a, **kw)
        by_name = {Path(f.path).name: f for f in infos}
        return [by_name[n] for n in self._order]

    def __getattr__(self, name):
        return getattr(self._real, name)


class Test1LatestFileByModificationTime(unittest.TestCase):
    def test_most_recently_modified_file_is_selected(self):
        with TempDir() as d:
            older = d / 'a.xlsx'
            newer = d / 'b.xlsx'
            _write_xlsx(older)
            _write_xlsx(newer)
            base_time = time.time()
            os.utime(older, (base_time, base_time))
            os.utime(newer, (base_time + 10, base_time + 10))

            # older listed first -- a selection that just picked
            # file_infos[0] would pick the wrong (older) one here.
            access = _FixedOrderAccess(['a.xlsx', 'b.xlsx'])
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=access)
            self.assertEqual(resource.file_path, str(newer))


class Test2DeterministicTieBreak(unittest.TestCase):
    def test_equal_modification_times_break_tie_by_path(self):
        with TempDir() as d:
            a = d / 'a.xlsx'
            z = d / 'z.xlsx'
            _write_xlsx(a)
            _write_xlsx(z)
            same_time = time.time()
            os.utime(a, (same_time, same_time))
            os.utime(z, (same_time, same_time))

            # a.xlsx listed first -- max()'s tie-break key is
            # (mtime, path), so the lexicographically larger path (z)
            # must win regardless of which one was seen first.
            access = _FixedOrderAccess(['a.xlsx', 'z.xlsx'])
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=access)
            self.assertEqual(resource.file_path, str(z))

    def test_tie_break_is_stable_regardless_of_input_order(self):
        # Same fixture, opposite listing order -- confirms the result
        # doesn't depend on which order the files happen to be seen in.
        with TempDir() as d:
            z = d / 'z.xlsx'
            a = d / 'a.xlsx'
            _write_xlsx(z)
            _write_xlsx(a)
            same_time = time.time()
            os.utime(a, (same_time, same_time))
            os.utime(z, (same_time, same_time))

            access = _FixedOrderAccess(['z.xlsx', 'a.xlsx'])
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=access)
            self.assertEqual(resource.file_path, str(z))


class Test3FileSetOrderIndependentOfDiscoveryOrder(unittest.TestCase):
    def test_sorted_output_regardless_of_scrambled_input_order(self):
        with TempDir() as d:
            for name in ('c.xlsx', 'a.xlsx', 'b.xlsx'):
                _write(d / name)

            class ScramblingAccess(FileAccessImpl):
                """Wraps select_file_infos to hand results back in a
                deliberately different order each call, simulating
                unstable filesystem enumeration order."""

                def __init__(self, order):
                    super().__init__()
                    self._order = order

                def select_file_infos(self, *a, **kw):
                    infos = super().select_file_infos(*a, **kw)
                    by_name = {Path(f.path).name: f for f in infos}
                    return [by_name[n] for n in self._order]

            access1 = ScramblingAccess(['c.xlsx', 'a.xlsx', 'b.xlsx'])
            access2 = ScramblingAccess(['b.xlsx', 'c.xlsx', 'a.xlsx'])

            resource1 = build_file_set_resource(str(d), pattern='*.xlsx', source_access=access1)
            resource2 = build_file_set_resource(str(d), pattern='*.xlsx', source_access=access2)

            names1 = [Path(f.path).name for f in resource1.files]
            names2 = [Path(f.path).name for f in resource2.files]
            self.assertEqual(names1, names2, 'file-set order depended on discovery order')
            self.assertEqual(names1, ['a.xlsx', 'b.xlsx', 'c.xlsx'])


class Test4FingerprintChangesOnAddOrRemove(unittest.TestCase):
    def test_adding_a_file_changes_the_fingerprint(self):
        with TempDir() as d:
            _write(d / 'a.xlsx')
            resource1 = build_file_set_resource(str(d), pattern='*.xlsx', source_access=tc.LOCAL_FILE_ACCESS)
            sig1 = resource1.source_fingerprint('files').source_signature

            _write(d / 'b.xlsx')
            resource2 = build_file_set_resource(str(d), pattern='*.xlsx', source_access=tc.LOCAL_FILE_ACCESS)
            sig2 = resource2.source_fingerprint('files').source_signature

            self.assertNotEqual(sig1, sig2)

    def test_removing_a_file_changes_the_fingerprint(self):
        with TempDir() as d:
            _write(d / 'a.xlsx')
            _write(d / 'b.xlsx')
            resource1 = build_file_set_resource(str(d), pattern='*.xlsx', source_access=tc.LOCAL_FILE_ACCESS)
            sig1 = resource1.source_fingerprint('files').source_signature

            (d / 'b.xlsx').unlink()
            resource2 = build_file_set_resource(str(d), pattern='*.xlsx', source_access=tc.LOCAL_FILE_ACCESS)
            sig2 = resource2.source_fingerprint('files').source_signature

            self.assertNotEqual(sig1, sig2)


class Test5FingerprintChangesOnSizeOrMtime(unittest.TestCase):
    def test_changed_file_size_changes_the_fingerprint(self):
        with TempDir() as d:
            _write(d / 'a.xlsx', b'short')
            resource1 = build_file_set_resource(str(d), pattern='*.xlsx', source_access=tc.LOCAL_FILE_ACCESS)
            sig1 = resource1.source_fingerprint('files').source_signature

            _write(d / 'a.xlsx', b'a much longer piece of content than before')
            resource2 = build_file_set_resource(str(d), pattern='*.xlsx', source_access=tc.LOCAL_FILE_ACCESS)
            sig2 = resource2.source_fingerprint('files').source_signature

            self.assertNotEqual(sig1, sig2)

    def test_changed_modification_time_changes_the_fingerprint(self):
        with TempDir() as d:
            path = d / 'a.xlsx'
            _write(path, b'same content throughout')
            resource1 = build_file_set_resource(str(d), pattern='*.xlsx', source_access=tc.LOCAL_FILE_ACCESS)
            sig1 = resource1.source_fingerprint('files').source_signature

            new_time = time.time() + 3600
            os.utime(path, (new_time, new_time))
            resource2 = build_file_set_resource(str(d), pattern='*.xlsx', source_access=tc.LOCAL_FILE_ACCESS)
            sig2 = resource2.source_fingerprint('files').source_signature

            self.assertNotEqual(sig1, sig2, 'same content, only mtime changed, but fingerprint stayed the same')


class Test6FingerprintStableAcrossDiscoveryOrder(unittest.TestCase):
    def test_fingerprint_signature_identical_regardless_of_scrambled_input_order(self):
        with TempDir() as d:
            for name in ('c.xlsx', 'a.xlsx', 'b.xlsx'):
                _write(d / name)

            class ScramblingAccess(FileAccessImpl):
                def __init__(self, order):
                    super().__init__()
                    self._order = order

                def select_file_infos(self, *a, **kw):
                    infos = super().select_file_infos(*a, **kw)
                    by_name = {Path(f.path).name: f for f in infos}
                    return [by_name[n] for n in self._order]

            resource1 = build_file_set_resource(
                str(d), pattern='*.xlsx', source_access=ScramblingAccess(['c.xlsx', 'a.xlsx', 'b.xlsx']),
            )
            resource2 = build_file_set_resource(
                str(d), pattern='*.xlsx', source_access=ScramblingAccess(['b.xlsx', 'c.xlsx', 'a.xlsx']),
            )

            sig1 = resource1.source_fingerprint('files').source_signature
            sig2 = resource2.source_fingerprint('files').source_signature
            self.assertEqual(sig1, sig2, 'fingerprint signature depended on discovery order')


class Test7FingerprintUsesAlreadySelectedFiles(unittest.TestCase):
    def test_file_set_resource_lists_the_directory_exactly_once(self):
        with TempDir() as d:
            _write(d / 'a.xlsx')
            _write(d / 'b.xlsx')

            call_count = [0]

            class CountingAccess(FileAccessImpl):
                def select_file_infos(self, *a, **kw):
                    call_count[0] += 1
                    return super().select_file_infos(*a, **kw)

            resource = build_file_set_resource(str(d), pattern='*.xlsx', source_access=CountingAccess())
            self.assertEqual(call_count[0], 1)

            # Using both .files and .source_fingerprint() afterward must
            # not trigger a second listing -- both read the same,
            # already-selected tuple.
            _ = resource.files
            _ = resource.source_fingerprint('files')
            self.assertEqual(call_count[0], 1, 'accessing .files/.source_fingerprint() re-listed the directory')

    def test_latest_xlsx_resource_lists_the_directory_exactly_once(self):
        with TempDir() as d:
            _write_xlsx(d / 'a.xlsx')

            call_count = [0]

            class CountingAccess(FileAccessImpl):
                def select_file_infos(self, *a, **kw):
                    call_count[0] += 1
                    return super().select_file_infos(*a, **kw)

            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=CountingAccess())
            self.assertEqual(call_count[0], 1)

            _ = resource.source_fingerprint('files')
            self.assertEqual(call_count[0], 1, 'source_fingerprint() re-listed the directory instead of reusing the selection')


class Test8NoMatchAndOnEmpty(unittest.TestCase):
    def test_no_matching_files_raises_by_default(self):
        with TempDir() as d:
            with self.assertRaises(NoMatchingFilesError):
                build_file_set_resource(str(d), pattern='*.xlsx', source_access=tc.LOCAL_FILE_ACCESS)

    def test_on_empty_empty_returns_a_valid_empty_resource_instead(self):
        with TempDir() as d:
            resource = build_file_set_resource(
                str(d), pattern='*.xlsx', source_access=tc.LOCAL_FILE_ACCESS, on_empty='empty',
            )
            self.assertEqual(resource.files, ())

    def test_latest_xlsx_no_match_raises_clearly(self):
        with TempDir() as d:
            with self.assertRaises(NoMatchingFilesError):
                build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=tc.LOCAL_FILE_ACCESS)


class Test9DirectFilePathRejectedByFolderScanning(unittest.TestCase):
    """Found by external review, confirmed directly before fixing:
    passing a direct file path to select_file_infos() (both local and
    SMB branches) bypassed every filter entirely -- pattern, hidden,
    system, temp, min_age_seconds -- since it just returned that one file
    immediately without ever reaching the filtering loop. A caller
    scanning for '*.xlsx' would get back a '.txt' file it explicitly
    named, with no error anywhere. select_fixed_file() already exists as
    the correct, dedicated API for "I know the exact file" -- confirmed
    no real task in this project relies on the old, bypassing behavior
    before making this change."""

    def test_direct_file_path_is_rejected_not_silently_selected(self):
        with TempDir() as d:
            path = d / 'x.txt'
            _write(path, b'not an xlsx')

            with self.assertRaises(ValueError):
                tc.LOCAL_FILE_ACCESS.select_file_infos(str(path), pattern='*.xlsx')

    def test_select_fixed_file_still_works_correctly_for_this_case(self):
        with TempDir() as d:
            path = d / 'x.txt'
            _write(path, b'not an xlsx')

            result = tc.LOCAL_FILE_ACCESS.select_fixed_file(str(path))
            self.assertEqual(result, str(path))

    def test_folder_scanning_still_works_normally(self):
        with TempDir() as d:
            _write(d / 'a.xlsx')

            result = tc.LOCAL_FILE_ACCESS.select_file_infos(str(d), pattern='*.xlsx')
            self.assertEqual(len(result), 1)


class Test10GetSheetRawRowsIsGenuinelyHeaderless(unittest.TestCase):
    """Found by external review: get_sheet_rows()/get_range()'s header=
    argument was a confirmed no-op -- [rows[0], *rows[1:]] is
    mathematically identical to rows, so header=True/False produced
    identical results. Not fixed in place, since hr_task.py's own
    sheet_to_raw_dataframe() already depended on this exact, documented
    behavior. get_sheet_raw_rows() is the new, genuinely header-agnostic
    method -- migrated hr_task.py to it, removing its own etl.header()/
    etl.data() reconstruction workaround entirely, confirmed unchanged
    by the full hr_task.py regression."""

    def test_row_0_is_genuinely_included_as_data(self):
        with TempDir() as d:
            path = d / 'x.xlsx'
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(['col-a', 'col-b'])
            ws.append([1, 2])
            wb.save(path)

            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=tc.LOCAL_FILE_ACCESS)
            raw = resource.get_sheet_raw_rows('Sheet')
            self.assertEqual(raw, [('col-a', 'col-b'), (1, 2)])
            resource.close()


class Test11XlsxInfoDoesNotSwallowCloseFailures(unittest.TestCase):
    """Found by external review: xlsx_info() explicitly caught and
    discarded any failure from wb.close() (`except Exception: pass`),
    which didn't follow the same "successful work must not silently hide
    cleanup failure" policy applied everywhere else in this project. The
    underlying binary stream still gets its own, separate cleanup via the
    surrounding `with self.open_binary(...)` regardless of whether
    wb.close() itself succeeds, so nothing is lost by letting this
    propagate naturally."""

    def test_close_failure_propagates_instead_of_being_swallowed(self):
        with TempDir() as d:
            path = d / 'x.xlsx'
            _write_xlsx(path)

            import openpyxl
            original_close = openpyxl.Workbook.close

            def failing_close(self):
                raise OSError('workbook close failed')

            openpyxl.Workbook.close = failing_close
            try:
                with self.assertRaises(OSError):
                    tc.LOCAL_FILE_ACCESS.xlsx_info(str(path))
            finally:
                openpyxl.Workbook.close = original_close

    def test_del_wb_runs_before_gc_collect_even_when_close_fails(self):
        # Found by a further review, confirmed directly before fixing:
        # if wb.close() itself raised, `del wb` right after it in the
        # same finally: block was skipped entirely -- an exception from
        # one statement in a finally: block skips the rest of that same
        # block, a general Python semantic confirmed directly in
        # isolation first, before assuming it applied here too. That
        # left wb as a live, reachable local variable at the point the
        # outer gc.collect() ran, weakening the exact stuck-handle
        # workaround it exists to guarantee, on the one path (a failing
        # close()) where that guarantee matters most.
        #
        # Verified here via direct frame introspection at gc.collect()'s
        # own call site -- the unambiguous signal for this specific
        # mechanism. A weakref-based check would not have been reliable:
        # a real openpyxl Workbook likely has its own internal reference
        # cycles (confirmed directly this project needed gc.collect() at
        # all specifically because ordinary refcounting alone wasn't
        # enough), so the object could still be alive for an unrelated
        # reason regardless of whether `del wb` specifically ran.
        import sys
        fa_module = sys.modules['task_core.file_access']

        with TempDir() as d:
            path = d / 'x.xlsx'
            _write_xlsx(path)

            observations = []
            original_gc_collect = fa_module.gc.collect

            def spy_collect(*a, **k):
                caller_frame = sys._getframe(1)
                observations.append('wb' in caller_frame.f_locals)
                return original_gc_collect(*a, **k)

            import openpyxl
            original_close = openpyxl.Workbook.close

            def failing_close(self):
                raise RuntimeError('close failed')

            fa_module.gc.collect = spy_collect
            openpyxl.Workbook.close = failing_close
            try:
                with self.assertRaises(RuntimeError):
                    tc.LOCAL_FILE_ACCESS.xlsx_info(str(path))
            finally:
                openpyxl.Workbook.close = original_close
                fa_module.gc.collect = original_gc_collect

            self.assertEqual(len(observations), 1, 'gc.collect() should have run exactly once')
            self.assertFalse(
                observations[0],
                "wb was still a live local variable when gc.collect() ran -- del wb did not run before it",
            )

    def test_normal_xlsx_info_still_works(self):
        with TempDir() as d:
            path = d / 'x.xlsx'
            _write_xlsx(path)
            sheets, tables = tc.LOCAL_FILE_ACCESS.xlsx_info(str(path))
            self.assertEqual(sheets, ['Sheet'])


class _CountingXlsxInfoAccess(FileAccessImpl):
    """Wraps the real xlsx_info() with a call counter, delegating to the
    genuine implementation -- not a fake that could silently drift from
    real behavior, the same real file_access class every other test in
    this file already exercises."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.xlsx_info_calls = 0

    def xlsx_info(self, *args, **kwargs):
        self.xlsx_info_calls += 1
        return super().xlsx_info(*args, **kwargs)


class Test12ExcelMetadataIsGenuinelyLazy(unittest.TestCase):
    """Found during an optimization review: build_excel_resource() called
    xlsx_info() (tables=True, its own default -- the full, expensive
    scan) unconditionally, before any skip decision could be made.
    Confirmed directly beforehand that collect_source_fingerprints()
    deliberately reuses get_resource() rather than a throwaway instance
    (so a task that does run doesn't reopen the workbook), which means
    that eager scan ran during every fingerprint collection too -- even
    for a task about to be skipped because its source hadn't changed,
    and even though source_fingerprint() (confirmed directly) never
    reads .sheets or .tables at all, only pre-captured selection
    metadata."""

    def test_building_the_resource_does_not_open_the_workbook_at_all(self):
        with TempDir() as d:
            path = d / 'x.xlsx'
            _write_xlsx(path)
            access = _CountingXlsxInfoAccess()
            build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=access)
            self.assertEqual(
                access.xlsx_info_calls, 0,
                'xlsx_info() ran during resource construction, before .sheets/.tables was ever accessed',
            )

    def test_fingerprinting_alone_never_opens_the_workbook(self):
        # The actual scenario this fix targets: a task_context collecting
        # fingerprints to decide whether to skip, never touching .sheets
        # or .tables at all.
        with TempDir() as d:
            path = d / 'x.xlsx'
            _write_xlsx(path)
            access = _CountingXlsxInfoAccess()
            ctx = tc.task_context(
                task_name='t',
                loaders={'r': lambda: build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=access)},
            )
            ctx.tracked_sources = [tc.TrackedResourceSource('r')]
            ctx.collect_source_fingerprints()
            self.assertEqual(
                access.xlsx_info_calls, 0,
                'fingerprint collection opened the workbook for metadata it never needed',
            )

    def test_accessing_sheets_triggers_exactly_one_load(self):
        with TempDir() as d:
            path = d / 'x.xlsx'
            _write_xlsx(path)
            access = _CountingXlsxInfoAccess()
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=access)
            self.assertEqual(access.xlsx_info_calls, 0)
            sheets = resource.sheets
            self.assertEqual(access.xlsx_info_calls, 1)
            self.assertEqual(sheets, ['Sheet'])

    def test_sheets_and_tables_share_one_load_not_two(self):
        # The design decision made explicitly during this discussion:
        # one shared lazy load between .sheets and .tables, not two
        # independent ones -- confirmed real tasks (hr_task.py) access
        # .sheets without .tables, but real business cases also need
        # both from the same source file, and splitting them would mean
        # a second, separate workbook open whichever is accessed second.
        with TempDir() as d:
            path = d / 'x.xlsx'
            _write_xlsx(path)
            access = _CountingXlsxInfoAccess()
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=access)

            _ = resource.sheets
            self.assertEqual(access.xlsx_info_calls, 1)
            _ = resource.tables  # must NOT trigger a second load
            self.assertEqual(
                access.xlsx_info_calls, 1,
                'accessing .tables after .sheets triggered a second, separate workbook open',
            )

            # and the reverse order
            access2 = _CountingXlsxInfoAccess()
            resource2 = build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=access2)
            _ = resource2.tables
            self.assertEqual(access2.xlsx_info_calls, 1)
            _ = resource2.sheets
            self.assertEqual(access2.xlsx_info_calls, 1)

    def test_repeated_access_does_not_reload(self):
        with TempDir() as d:
            path = d / 'x.xlsx'
            _write_xlsx(path)
            access = _CountingXlsxInfoAccess()
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=access)
            _ = resource.sheets
            _ = resource.tables
            _ = resource.sheets
            _ = resource.tables
            self.assertEqual(access.xlsx_info_calls, 1)

    def test_get_table_still_works_via_the_lazy_tables_property(self):
        with TempDir() as d:
            path = d / 'x.xlsx'
            from openpyxl import Workbook
            from openpyxl.worksheet.table import Table
            wb = Workbook()
            ws = wb.active
            ws.append(['col_a', 'col_b'])
            ws.append([1, 2])
            ws.add_table(Table(displayName='MyTable', ref='A1:B2'))
            wb.save(path)

            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx')
            tbl = resource.get_table('MyTable')
            rows = list(tc.etl.data(tbl)) if hasattr(tc, 'etl') else list(__import__('petl').data(tbl))
            self.assertEqual(rows, [(1, 2)])


class _RejectOverlappingOpensAccess(FileAccessImpl):
    """Rejects xlsx_info() while a retained workbook (opened via
    open_workbook()) is still active -- the same event-order check the
    external review that found this bug suggested, not merely that
    get_table() works on a local file."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._workbook_open_active = False

    def open_workbook(self, *args, **kwargs):
        self._workbook_open_active = True
        cm = super().open_workbook(*args, **kwargs)
        outer = self

        class _Wrapper:
            def __enter__(self_w):
                return cm.__enter__()
            def __exit__(self_w, *exc):
                outer._workbook_open_active = False
                return cm.__exit__(*exc)
        return _Wrapper()

    def xlsx_info(self, *args, **kwargs):
        if self._workbook_open_active:
            raise RuntimeError(
                'xlsx_info() called while a retained workbook is still open -- overlapping access'
            )
        return super().xlsx_info(*args, **kwargs)


class Test13GetTableDoesNotOverlapWithRetainedWorkbook(unittest.TestCase):
    """Found by external review, confirmed directly before fixing:
    get_table() called self._ensure_workbook() -- opening and retaining
    the main workbook for the resource's whole lifetime -- before
    evaluating self.tables[name], which on first access triggers
    xlsx_info()'s own, completely separate workbook open. That meant two
    workbooks were briefly open on the same file simultaneously, where
    sequential access (metadata loads, opens, and closes fully, only
    then the retained workbook opens) is what this project's own file
    access layer is otherwise careful to guarantee, especially over
    SMB/DFS, where handle behavior has already been directly, personally
    confirmed unreliable."""

    def _build_xlsx_with_table(self, d):
        from openpyxl import Workbook
        from openpyxl.worksheet.table import Table as XlsxTable
        path = d / 'x.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.append(['col_a', 'col_b'])
        ws.append([1, 2])
        ws.add_table(XlsxTable(displayName='MyTable', ref='A1:B2'))
        wb.save(path)
        return path

    def test_get_table_does_not_overlap(self):
        with TempDir() as d:
            self._build_xlsx_with_table(d)
            access = _RejectOverlappingOpensAccess()
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=access)
            tbl = resource.get_table('MyTable')  # must not raise
            self.assertEqual(list(__import__('petl').data(tbl)), [(1, 2)])

    def test_get_map_does_not_overlap(self):
        with TempDir() as d:
            self._build_xlsx_with_table(d)
            access = _RejectOverlappingOpensAccess()
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=access)
            m = resource.get_map('MyTable')  # must not raise
            self.assertEqual(m, {1: 2})

    def test_get_sheet_rows_then_get_table_does_not_overlap(self):
        # Found by a further review, confirmed directly before fixing:
        # the earlier fix only reordered get_table()'s own two internal
        # steps -- it didn't protect against _ensure_workbook() already
        # having been triggered by an EARLIER, different method call.
        # get_sheet_rows() retains the workbook first here; get_table()
        # then triggers .tables' own, separate xlsx_info() open on top
        # of it.
        with TempDir() as d:
            self._build_xlsx_with_table(d)
            access = _RejectOverlappingOpensAccess()
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=access)
            resource.get_sheet_rows('Sheet')
            resource.get_table('MyTable')  # must not raise

    def test_get_sheet_raw_rows_then_get_map_does_not_overlap(self):
        with TempDir() as d:
            self._build_xlsx_with_table(d)
            access = _RejectOverlappingOpensAccess()
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=access)
            resource.get_sheet_raw_rows('Sheet')
            resource.get_map('MyTable')  # must not raise

    def test_get_range_then_tables_property_does_not_overlap(self):
        with TempDir() as d:
            self._build_xlsx_with_table(d)
            access = _RejectOverlappingOpensAccess()
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=access)
            resource.get_range('Sheet', 'A1:B2')
            _ = resource.tables  # must not raise


class Test14ExcelResourceConstructorStaysPubliclyCompatible(unittest.TestCase):
    """Found by external review: excel_resource is genuinely exported
    from task_core's own public facade (confirmed directly --
    `from task_core import excel_resource` works), so making sheets/
    tables lazy shouldn't have silently required them to disappear from
    the constructor's own signature entirely. The pre-existing
    constructor accepted both as required positional args; this
    confirms a caller providing them directly still works exactly as it
    did before, and short-circuits the lazy load entirely rather than
    silently discarding what was explicitly given."""

    def test_old_style_call_with_explicit_sheets_and_tables_still_works(self):
        from task_core.resources.excel import excel_resource
        # The review's own exact reproduction of the regression.
        resource = excel_resource('x.xlsx', ['Sheet'], {})
        self.assertEqual(resource.sheets, ['Sheet'])
        self.assertEqual(resource.tables, {})

    def test_providing_sheets_and_tables_short_circuits_the_lazy_load(self):
        from task_core.resources.excel import excel_resource

        class _RaisesIfCalled:
            def xlsx_info(self, *args, **kwargs):
                raise AssertionError('xlsx_info() should never be called -- metadata was already provided')

        resource = excel_resource(
            'x.xlsx', ['SheetA', 'SheetB'], {'T1': {'sheet': 'SheetA', 'range_string': 'A1:B2'}},
            source_access=_RaisesIfCalled(),
        )
        self.assertEqual(resource.sheets, ['SheetA', 'SheetB'])  # must not raise
        self.assertEqual(resource.tables, {'T1': {'sheet': 'SheetA', 'range_string': 'A1:B2'}})

    def test_neither_provided_still_stays_genuinely_lazy(self):
        from task_core.resources.excel import excel_resource
        resource = excel_resource('x.xlsx')
        self.assertIsNone(resource._sheets_value)
        self.assertIsNone(resource._tables_value)


class Test15RowMetadataIsCachedByKey(unittest.TestCase):
    """Found on the original optimization proposal's "smaller
    improvements" list: read_excel_row_metadata() opens its own,
    separate handle on every call, independent of whatever the resource
    already has open -- a genuine, real cost the resource represents
    one immutable selected workbook, so caching it is always safe, not
    an optimization that could introduce staleness.

    Also owns the defensive-copy property for both resources
    (returns_an_independent_copy_* below): read_excel_row_metadata()
    cached the raw {row_number: value} dict and returned it directly, so a
    caller mutating what it got back silently corrupted the cached answer
    for a later call with the same key -- the same class of regression
    already fixed for get_sheet_raw_rows() (Test17). A separate class
    covered the identical three properties for a while, differing only in
    which key it mutated; removed as duplication rather than kept as a
    second, drifting copy."""

    def _build_xlsx(self, d):
        from openpyxl import Workbook
        path = d / 'x.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.append(['a', 'b'])
        ws.append([1, 2])
        ws.row_dimensions[2].outline_level = 1
        wb.save(path)
        return path

    def test_excel_resource_second_call_does_not_reopen(self):
        with TempDir() as d:
            self._build_xlsx(d)
            access = _CountingOpenBinaryAccess()
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=access)

            m1 = resource.read_excel_row_metadata(sheet=0, mode='outline')
            calls_after_first = access.open_binary_calls
            m2 = resource.read_excel_row_metadata(sheet=0, mode='outline')
            calls_after_second = access.open_binary_calls

            self.assertEqual(calls_after_first, calls_after_second, 'a second, identical call reopened a handle')
            self.assertEqual(m1, m2)

    def test_excel_resource_different_key_does_reopen(self):
        with TempDir() as d:
            self._build_xlsx(d)
            access = _CountingOpenBinaryAccess()
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx', source_access=access)

            resource.read_excel_row_metadata(sheet=0, mode='outline')
            calls_after_first = access.open_binary_calls
            resource.read_excel_row_metadata(sheet=0, mode='indent', column=0)
            calls_after_second = access.open_binary_calls

            self.assertGreater(calls_after_second, calls_after_first, 'a genuinely different key should not share a cache entry')

    def test_file_set_resource_second_call_does_not_reopen(self):
        with TempDir() as d:
            self._build_xlsx(d)
            access = _CountingOpenBinaryAccess()
            resource = build_file_set_resource(str(d), pattern='*.xlsx', source_access=access)
            selected_file = resource.files[0]

            m1 = resource.read_excel_row_metadata(selected_file, sheet=0, mode='outline')
            calls_after_first = access.open_binary_calls
            m2 = resource.read_excel_row_metadata(selected_file, sheet=0, mode='outline')
            calls_after_second = access.open_binary_calls

            self.assertEqual(calls_after_first, calls_after_second)
            self.assertEqual(m1, m2)

    def test_excel_resource_returns_an_independent_copy_not_the_internal_cache(self):
        # Found by a further review, confirmed directly before fixing:
        # a caller mutating what read_excel_row_metadata() returned
        # silently corrupted the cached answer for a later call with
        # the same key -- the same class of regression already fixed
        # for get_sheet_raw_rows().
        with TempDir() as d:
            self._build_xlsx(d)
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx')

            first = resource.read_excel_row_metadata(sheet=0, mode='outline')
            first[2] = 99
            second = resource.read_excel_row_metadata(sheet=0, mode='outline')

            self.assertNotEqual(second.get(2), 99, "mutating the first call's result corrupted the cached answer")
            self.assertIsNot(first, second)

    def test_file_set_resource_returns_an_independent_copy_not_the_internal_cache(self):
        with TempDir() as d:
            self._build_xlsx(d)
            resource = build_file_set_resource(str(d), pattern='*.xlsx')
            selected_file = resource.files[0]

            first = resource.read_excel_row_metadata(selected_file, sheet=0, mode='outline')
            first[2] = 99
            second = resource.read_excel_row_metadata(selected_file, sheet=0, mode='outline')

            self.assertNotEqual(second.get(2), 99)
            self.assertIsNot(first, second)

    def test_close_clears_the_row_metadata_cache(self):
        # Found by a further review, confirmed directly: close() cleared
        # every other data cache (_table_cache, _map_cache, _range_cache,
        # _sheet_cache, _raw_cache) but genuinely omitted this one.
        with TempDir() as d:
            self._build_xlsx(d)
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx')
            resource.read_excel_row_metadata(sheet=0, mode='outline')
            self.assertTrue(resource._row_metadata_cache)

            resource.close()

            self.assertEqual(resource._row_metadata_cache, {})


class _CountingOpenBinaryAccess(FileAccessImpl):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.open_binary_calls = 0

    def open_binary(self, *args, **kwargs):
        self.open_binary_calls += 1
        return super().open_binary(*args, **kwargs)


class Test16SheetRowsSharesRawRowsMaterialization(unittest.TestCase):
    """Found on the original optimization proposal's "smaller
    improvements" list: get_sheet_rows() and get_sheet_raw_rows() both
    independently called the identical list(ws.values), into two
    separate caches -- genuinely reading and storing the same sheet's
    rows twice if both were ever used for the same sheet."""

    def _build_xlsx(self, d):
        from openpyxl import Workbook
        path = d / 'x.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.append(['a', 'b'])
        ws.append([1, 2])
        ws.append([3, 4])
        wb.save(path)
        return path

    def test_get_sheet_rows_alone_populates_the_shared_raw_cache(self):
        with TempDir() as d:
            self._build_xlsx(d)
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx')
            resource.get_sheet_rows('Sheet')
            self.assertIn('Sheet', resource._raw_cache, 'get_sheet_rows() did not share get_sheet_raw_rows()\'s own cache')

    def test_both_return_correct_and_consistent_values(self):
        with TempDir() as d:
            self._build_xlsx(d)
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx')
            tbl = resource.get_sheet_rows('Sheet')
            raw = resource.get_sheet_raw_rows('Sheet')

            self.assertEqual(list(etl.data(tbl)), [(1, 2), (3, 4)])
            self.assertEqual(raw, [('a', 'b'), (1, 2), (3, 4)])

    def test_get_sheet_raw_rows_returns_an_independent_copy_not_the_internal_cache(self):
        # Updated: an earlier version of this test asserted the OPPOSITE
        # -- that get_sheet_raw_rows() returned the exact same object as
        # the internal cache directly. That was itself the bug a further
        # review found (see Test17 below): mutating what it returned
        # silently changed get_sheet_rows()'s own, already-cached table
        # too, since both were built from that one, shared object.
        with TempDir() as d:
            self._build_xlsx(d)
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx')
            resource.get_sheet_rows('Sheet')
            raw = resource.get_sheet_raw_rows('Sheet')
            self.assertIsNot(
                resource._raw_cache['Sheet'], raw,
                'get_sheet_raw_rows() returned the internal cache object directly, not an independent copy',
            )
            self.assertEqual(resource._raw_cache['Sheet'], raw, 'the copy should still be equal in content')

    def test_repeated_get_sheet_rows_calls_return_the_same_object(self):
        with TempDir() as d:
            self._build_xlsx(d)
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx')
            a = resource.get_sheet_rows('Sheet')
            b = resource.get_sheet_rows('Sheet')
            self.assertIs(a, b)

    def test_header_false_deprecation_warning_still_fires(self):
        with TempDir() as d:
            self._build_xlsx(d)
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx')
            with self.assertWarns(DeprecationWarning):
                resource.get_sheet_rows('Sheet', header=False)


class Test17GetSheetRawRowsReturnsAnIndependentCopy(unittest.TestCase):
    """Found by a further review, confirmed directly before fixing:
    after get_sheet_rows()/get_sheet_raw_rows() were made to share one
    materialization (Test16 above), mutating what get_sheet_raw_rows()
    returned silently changed get_sheet_rows()'s own, already-cached
    petl table too -- both were built from the exact same, single list
    object. Previously, before that fix, the two methods had completely
    independent materializations, so this was never possible -- the
    shared-materialization fix introduced a real, if narrow, behavioral
    regression that hadn't been there before."""

    def _build_xlsx(self, d):
        from openpyxl import Workbook
        path = d / 'x.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.append(['a', 'b'])
        ws.append([1, 2])
        ws.append([3, 4])
        wb.save(path)
        return path

    def test_mutating_returned_raw_rows_does_not_affect_get_sheet_rows(self):
        # The reviewer's own exact reproduction.
        with TempDir() as d:
            self._build_xlsx(d)
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx')
            raw = resource.get_sheet_raw_rows('Sheet')
            table = resource.get_sheet_rows('Sheet')

            raw[1] = (9, 9)

            self.assertEqual(
                list(etl.data(table)), [(1, 2), (3, 4)],
                "mutating what get_sheet_raw_rows() returned silently changed get_sheet_rows()'s own table",
            )

    def test_two_separate_calls_return_independent_copies(self):
        with TempDir() as d:
            self._build_xlsx(d)
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx')
            raw_a = resource.get_sheet_raw_rows('Sheet')
            raw_b = resource.get_sheet_raw_rows('Sheet')

            raw_a[0] = (99, 99)

            self.assertNotEqual(raw_a, raw_b, 'two separate calls should not share a mutable object')
            self.assertIsNot(raw_a, raw_b)

    def test_the_underlying_read_still_only_happens_once(self):
        # The core benefit of the original shared-materialization fix
        # must survive this one -- confirmed via the internal cache
        # object's own identity staying constant across repeated calls,
        # not by counting a proxy metric (e.g. open_binary calls) that
        # doesn't actually isolate this specific cost from unrelated
        # ones (the metadata scan, the separate retained-workbook open).
        with TempDir() as d:
            self._build_xlsx(d)
            resource = build_latest_xlsx_resource(str(d), pattern='*.xlsx')

            resource.get_sheet_raw_rows('Sheet')
            internal_after_first = resource._raw_cache['Sheet']
            resource.get_sheet_raw_rows('Sheet')
            resource.get_sheet_rows('Sheet')
            resource.get_sheet_raw_rows('Sheet')

            self.assertIs(resource._raw_cache['Sheet'], internal_after_first)


class Test18RetainedWorkbookGetsTheSameGcTreatmentAsMetadata(unittest.TestCase):
    """xlsx_info() has run gc.collect() on every path for some time, on a
    confirmed production finding: openpyxl object cycles keep XLSX/ZIP
    handles open on SMB/DFS even after an explicit wb.close(), and
    refcounting alone does not release them. open_workbook() -- which
    serves the workbook excel_resource RETAINS for its entire lifetime,
    holding the remote stream open in SMB non-buffered mode -- had no
    gc.collect() on it at all. Confirmed directly by counting calls before
    fixing: xlsx_info() 1, open_workbook() 0, excel_resource.close() 0.

    The hardening had been applied where the bug was originally found
    rather than where the risk actually concentrates: xlsx_info()'s
    workbook lives for milliseconds, this one can live for a whole task.

    Fixing it needed two halves, not one. Adding gc.collect() to
    open_workbook() alone would have been defeated, because
    excel_resource.close() cleared self._wb only AFTER __exit__() returned
    -- confirmed directly with gc.get_referrers(), which listed the
    resource's own __dict__ as still holding the workbook at the instant
    the collect ran. That is the same way a skipped `del wb` defeats
    xlsx_info()'s own collect.
    """

    def _counting_access(self):
        """A source_access whose gc.collect() calls are counted, patched
        on the module.

        Reached through sys.modules because that is unambiguous. Until
        0.3.1 it was also NECESSARY: the facade re-exported a class named
        after its own module, so task_core.file_access resolved to the
        class and there was no way to reach the module by attribute. The
        class is now source_access and the attribute resolves to the
        module, but going through sys.modules costs nothing and states the
        intent."""
        import gc
        import sys

        module = sys.modules['task_core.file_access']
        counter = {'calls': 0}
        real_collect = module.gc.collect

        def counting_collect(*args, **kwargs):
            counter['calls'] += 1
            return real_collect(*args, **kwargs)

        module.gc.collect = counting_collect
        self.addCleanup(setattr, module.gc, 'collect', real_collect)
        return module.source_access(), counter

    def test_open_workbook_collects_on_the_success_path(self):
        with TempDir() as tmp:
            path = tmp / 'wb.xlsx'
            _write_xlsx(path)
            access, counter = self._counting_access()

            counter['calls'] = 0
            with access.open_workbook(str(path), read_only=True) as wb:
                self.assertTrue(wb.sheetnames)

            self.assertEqual(counter['calls'], 1)

    def test_open_workbook_collects_when_the_body_raises(self):
        # The path where it matters most and is hardest to notice: the
        # caller's attention is on the exception, not on whether cleanup ran.
        with TempDir() as tmp:
            path = tmp / 'wb.xlsx'
            _write_xlsx(path)
            access, counter = self._counting_access()

            counter['calls'] = 0
            with self.assertRaises(ValueError):
                with access.open_workbook(str(path), read_only=True):
                    raise ValueError('boom')

            self.assertEqual(counter['calls'], 1)

    def test_open_workbook_collects_and_drops_wb_even_if_close_raises(self):
        # del wb lives in its own inner finally: precisely so a raising
        # wb.close() cannot skip it. Verified by frame introspection at the
        # collect's own call site rather than by weakref -- a real openpyxl
        # Workbook has internal reference cycles (that being why
        # gc.collect() is needed at all), so a weakref could report the
        # object alive for an unrelated reason regardless of whether the
        # del actually ran.
        import gc
        import sys

        with TempDir() as tmp:
            path = tmp / 'wb.xlsx'
            _write_xlsx(path)

            module = sys.modules['task_core.file_access']
            real_collect = module.gc.collect
            seen = {}

            def introspecting_collect(*args, **kwargs):
                frame = sys._getframe(1)
                seen['is_open_workbook'] = frame.f_code.co_name == 'open_workbook'
                seen['wb_still_local'] = 'wb' in frame.f_locals
                return real_collect(*args, **kwargs)

            module.gc.collect = introspecting_collect
            self.addCleanup(setattr, module.gc, 'collect', real_collect)

            access = module.source_access()
            cm = access.open_workbook(str(path), read_only=True)
            wb = cm.__enter__()
            wb.close = lambda: (_ for _ in ()).throw(RuntimeError('close failed'))

            with self.assertRaises(RuntimeError):
                cm.__exit__(None, None, None)

            self.assertTrue(seen.get('is_open_workbook'))
            self.assertFalse(
                seen.get('wb_still_local'),
                'wb was still a live local in open_workbook() at gc.collect() time -- '
                'del wb was skipped by the raising wb.close()',
            )


class Test19ExcelResourceCloseDropsTheWorkbookBeforeCollecting(unittest.TestCase):
    """The second half of Test18's fix. excel_resource.close() now drops
    _wb and _workbook_cm BEFORE calling __exit__(), because
    open_workbook()'s gc.collect() runs from inside that call and a live
    self._wb at that moment defeats it entirely.

    Also brings close() to the same failure behavior excel_resource's
    caches already had and db_resource.close() separately gained: state
    cleared regardless, exception still propagated, second close a no-op.
    """

    def _resource(self, tmp):
        path = tmp / 'wb.xlsx'
        _write_xlsx(path)
        resource = tc.build_excel_resource(str(path))
        resource.get_sheet_raw_rows(resource.sheets[0])   # force the retained open
        self.assertIsNotNone(resource._wb)
        return resource

    def test_the_resource_no_longer_references_the_workbook_when_collect_runs(self):
        import sys

        with TempDir() as tmp:
            resource = self._resource(tmp)

            module = sys.modules['task_core.file_access']
            real_collect = module.gc.collect
            observed = {}

            def observing_collect(*args, **kwargs):
                observed['_wb'] = resource._wb
                observed['_workbook_cm'] = resource._workbook_cm
                return real_collect(*args, **kwargs)

            module.gc.collect = observing_collect
            self.addCleanup(setattr, module.gc, 'collect', real_collect)

            resource.close()

            self.assertIsNone(
                observed.get('_wb'),
                'excel_resource._wb still held the workbook at gc.collect() time',
            )
            self.assertIsNone(observed.get('_workbook_cm'))

    def test_a_failing_exit_still_clears_every_cache_and_still_raises(self):
        with TempDir() as tmp:
            resource = self._resource(tmp)
            resource.get_sheet_rows(resource.sheets[0])

            class FailingCm:
                def __exit__(self, *exc):
                    raise RuntimeError('workbook close failed')

            resource._workbook_cm = FailingCm()

            with self.assertRaises(RuntimeError):
                resource.close()

            self.assertIsNone(resource._wb)
            self.assertIsNone(resource._workbook_cm)
            for name in ('_table_cache', '_map_cache', '_range_cache',
                         '_sheet_cache', '_raw_cache', '_row_metadata_cache'):
                self.assertEqual(getattr(resource, name), {}, name)

    def test_a_second_close_after_a_failure_is_a_no_op(self):
        with TempDir() as tmp:
            resource = self._resource(tmp)
            exits = []

            class FailingCm:
                def __exit__(self, *exc):
                    exits.append(1)
                    raise RuntimeError('workbook close failed')

            resource._workbook_cm = FailingCm()

            with self.assertRaises(RuntimeError):
                resource.close()
            resource.close()   # must not retry the same failing context

            self.assertEqual(len(exits), 1)



if __name__ == '__main__':
    unittest.main()
