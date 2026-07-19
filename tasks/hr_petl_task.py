# -*- coding: utf-8 -*-
"""
hr_petl_task.py -- standalone example of the petl branch of task_core's
mixed-engine machinery (table_adapter left unset -> petl). Reads the same
real source folder/files as hr_task.py's ssch pipeline (a genuine pandas
migration from funnel_pandas.py), but is not itself a migration of anything
-- ssch2 is a deliberately petl-native rewrite of the same header/footer/
date-extraction problem, kept here as a working reference, not something
scheduled or depended on by hr_task.py.

Every real funnel_pandas.py migration uses table_adapter='pandas' -- this
file exists purely as an example of the other branch.
"""

from __future__ import annotations

import logging

from openpyxl import load_workbook

import petl as etl
from petl_util import *  # noqa: F403 -- also registers .table_skip()/.drop_blank_cols() onto petl's Table as a side effect, matching the real petl_util.py's own convention

from task_core import (
    PipelineSpec,
    ResourceEnvironment,
    SourceChangeCheckConfig,
    align_row_metadata,
    bind,
    build_resource_context,
    build_source_access,
    run_pipelines,
    setup_logging,
    xlsx_file_set,
)

try:
    from pgcreds import pgcreds as DEFAULT_PGCREDS
except ImportError:
    DEFAULT_PGCREDS = None


TASK_NAME = 'hr_petl_task'
BASE_PATH = r'\\telecom.local\dfs\Стратегическое развитие\pbi-data\HR'
PG_SCHEMA = 'bsr'
OUTPUT_EXCEL = True
OUTPUT_DB = True

log = logging.getLogger(TASK_NAME)

SOURCE_CHANGE_CHECK = SourceChangeCheckConfig(
    enabled=True,
    schema='bsr',
    table='task_scaffold_meta',
)


def read_ssch2_sheet(file_set, selected_file, *, sheet=0):
    # openpyxl directly, not pandas -- this pipeline is pandas-free
    # throughout. read_excel_row_metadata/align_row_metadata are
    # engine-neutral already; same SMB-readiness via file_set.open_file().
    metadata = file_set.read_excel_row_metadata(selected_file, sheet=sheet, mode='outline')

    with file_set.open_file(selected_file) as src:
        wb = load_workbook(src, read_only=True, data_only=True)
        ws = wb.worksheets[sheet] if isinstance(sheet, int) else wb[sheet]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]

    if len(metadata) != len(rows):
        log.warning(
            '%s: outline row count mismatch: %s rows from XML vs %s rows from the sheet '
            '-- some rows may get a None outline_level value',
            selected_file.relative_path, len(metadata), len(rows),
        )

    # first_row=1, not min(metadata): a genuinely untouched leading row
    # (no <row r="1"> in the raw XML at all) has no metadata entry, but
    # openpyxl's row-materialization still counts it as row 0 -- using
    # min(metadata) as first_row would then misalign every subsequent
    # row by one. Same fix as hr_task.py's read_ssch_sheet, verified
    # against the same construction there.
    first_row = 1
    aligned = align_row_metadata(metadata, first_row=first_row, n_rows=len(rows))

    n_cols = len(rows[0]) if rows else 0
    header = ['outline_level'] + [f'col{i}' for i in range(n_cols)]
    augmented = [[level, *row] for level, row in zip(aligned, rows)]
    return etl.wrap([header, *augmented])


class ssch2:
    # Petl-native rewrite of ssch's header/footer/date-extraction problem
    # -- same source files, genuinely different engine throughout, not a
    # pandas pipeline wrapped in etl.wrap() at the end.
    #
    # No hierarchy-building or lev.* columns, but the raw outline_level
    # value is kept as a plain column ('attr'), not discarded.
    spec = PipelineSpec(
        excel_name='ssch2.xlsx',
        db_table='hr_ssch2',
        # No db_contract needed -- columns are stable across files.
        # table_adapter left unset (petl) -- exercises the petl branch.
    )

    @staticmethod
    def process_file(file_set, selected_file):
        t = read_ssch2_sheet(file_set, selected_file, sheet=0)

        # Report month: table_skip(predicate) stops at (and includes) the
        # first "Период отчета" row; t2[1] is that row (tbl[0]=header,
        # tbl[1]=first data row). Real cell text is "... - DD.MM.YYYY".
        t2 = t.table_skip(
            lambda x: list(filter(
                lambda y: False if not isinstance(y, str) else y.startswith('Период отчета'), x,
            )) == []
        )
        date_cell = list(filter(
            lambda x: False if not isinstance(x, str) else x.startswith('Период отчета'), t2[1],
        ))[0]
        month = etl.dateparser('%d.%m.%Y')(date_cell.split(' - ')[1]).replace(day=1)

        # table_skip(predicate) finds and stops at the Организация anchor
        # row; .skip(1) promotes it to header; table_skip(3) skips the
        # same sub-header rows ssch does; rename(0, 'attr') names the
        # outline-level column (always position 0); drop_blank_cols
        # (pandas-internal roundtrip) removes empty columns without ever
        # passing a field spec to petl's strict cutout.
        tbl = (
            t
            .table_skip(lambda x: 'Организация' not in x).skip(1)
            .table_skip(3)
            .rename(0, 'attr')
            .drop_blank_cols()
            .cutout('Коэфф. текучести', 'Уволено')
            .select(lambda r: r['Организация'] is not None and r['Организация'] != 'Итого')
            .convert('Средняя численн.', float)
        )

        return tbl.addfield('Месяц', month)

    @classmethod
    def run(cls, ctx, *, source):
        tables = [cls.process_file(source, f) for f in source.files]
        return etl.cat(*tables)


SSCH_FILES = xlsx_file_set('ssch', pattern='*.xlsx', tracker=True)

RESOURCES = {
    'ssch_files': SSCH_FILES,
}


PIPELINES = {'ssch2': bind(ssch2, source=SSCH_FILES)}
RUN_SEQUENCE = ['ssch2']


def build_context(base_path=BASE_PATH, dfs_creds=None):
    source_access = build_source_access(dfs_creds=dfs_creds)
    env = ResourceEnvironment(base_path=base_path, file_access=source_access)
    return build_resource_context(TASK_NAME, RESOURCES, PIPELINES, RUN_SEQUENCE, env)


def main(base_path=BASE_PATH, output_creds=None, pg_schema=PG_SCHEMA, dfs_creds=None, force_run=False, smb_level=logging.WARNING):
    output_creds = DEFAULT_PGCREDS if output_creds is None else output_creds
    setup_logging(TASK_NAME, smb_level=smb_level)
    return run_pipelines(
        task_name=TASK_NAME,
        build_context=lambda: build_context(base_path, dfs_creds=dfs_creds),
        pipelines=PIPELINES,
        run_sequence=RUN_SEQUENCE,
        output_excel=OUTPUT_EXCEL,
        output_db=OUTPUT_DB,
        creds=output_creds,
        pg_schema=pg_schema,
        source_change_check=SOURCE_CHANGE_CHECK,
        force_run=force_run,
    )


if __name__ == '__main__':
    main()
